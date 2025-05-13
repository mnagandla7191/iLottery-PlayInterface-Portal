import os
from PageObjectLibrary import PageObject
from SeleniumLibrary import SeleniumLibrary
import openpyxl
import xlsxwriter
from robot.api.deco import keyword
from openpyxl import load_workbook
class Common_Base_Repo(PageObject):
        final_data=[]
        _locators = {
        
       
        "user_name": "id:loginModal-username",
        "password": "id:loginModal-password",
        }
        def Open_Browser_with_Portal_url(self,Browser_Type,url):
                print(Browser_Type)
                self.selib.open_browser(url,Browser_Type)

        
        @keyword('enter ${value}')
        def enter_values(self,onemorevalue,xpath,value):
                print(xpath)
                print(value)
                self.selib.wait_until_element_is_visible(xpath,30)
                self.selib.click_element(xpath)
                self.selib.input_text(xpath,value)

        @keyword('Clear ${field_value}')
        def Clear_text(self,onemorevalue,xpath):
                self.selib.wait_until_element_is_visible(xpath,30)
                self.selib.clear_element_text(xpath)

        @keyword('Click on ${Button_Name}')
        def Click_Button(self,Button_Name,XPATH):
                self.selib.wait_until_element_is_visible(XPATH,30)
                self.selib.click_element(XPATH)
                self.builtin.sleep(3)

        @keyword('Select value ${Drop_down_Value}')
        def Select_value_from_list(self,onemorevalue,select_location,value):
                print("Value is:",value)
                self.selib.wait_until_element_is_visible(select_location,30)
                self.selib.click_element(select_location)
                self.builtin.sleep(2)
                new_xpath = select_location+'/option[contains(text(),"'+value+'")]'
                
                print(new_xpath)
                self.selib.click_element(new_xpath)
        @keyword('Select Game ${Game_Name}')
        def Select_Game_from_the_list(self,onemorevalue,Game_Name,Bet_type,Draw_Name,flag=0):

                self.builtin.sleep(2)
                gname = self.convert_game_name_single_string(Game_Name)
                if flag==0:
                        xpath="//li/a[@class='nav-tile game-tile-jackpot game-tile-jackpot-"+gname+"']"
                else:
                        xpath="//li[contains(@class,'favorites')]/a[@data-gamename='"+gname+"']"
                self.selib.wait_until_element_is_visible(xpath,30)
                self.selib.click_element(xpath)
                self.builtin.sleep(3)
                if('Pick 3' in Game_Name or 'Pick 4' in Game_Name):
                        print("bet type is:",Bet_type)
                        print("Draw name is:",Draw_Name)
                        xpath_betType="//select[@id='betType']"
                        xpath_Drawname="//select[@id='drawNames']"
                        self.Select_value_from_list(onemorevalue,xpath_betType,Bet_type)
                        self.Select_value_from_list(onemorevalue,xpath_Drawname,Draw_Name)
                if("All or Nothing"in Game_Name):
                        xpath_Drawname="//select[@id='drawNames']"
                        self.Select_value_from_list(onemorevalue,xpath_Drawname,Draw_Name)

        
        @keyword('Select check box ${Checkbox}')
        def Agree_available_option(self,onemorevalue,select_location):
                self.selib.wait_until_element_is_visible(select_location,30)
                check_status = self.selib.get_element_attribute(select_location,'class')
                if('has-error' in check_status):
                        self.selib.click_element(select_location)
                self.builtin.sleep(2)
        @keyword('Accept ${Checkbox}')
        def Accept_available_option(self,onemorevalue,select_location):
                self.selib.wait_until_element_is_visible(select_location,30)
                self.selib.click_element(select_location)
                self.builtin.sleep(2)
        
        @keyword('Uncheck ${un_checkbox}')
        def uncheck_available_option(self,onemorevalue,select_location):
                self.selib.wait_until_element_is_visible(select_location,30)
                check_status = self.selib.get_element_attribute(select_location,'class')
                print(check_status)
                if('has-error' not in check_status):
                        self.selib.click_element(select_location)
                self.builtin.sleep(2)



        @keyword('Get Error Message ${Message}')
        def Get_Error_Message_displayed_in_Portal_page(self,onemorevalue,message_location):
             self.selib.wait_until_element_is_visible(message_location,30)
             error_message =  self.selib.get_text(message_location)
             return     error_message
        
        @keyword('Get text Message ${Message}')
        def Get_test_from_Portal_page(self,onemorevalue,message_location):
             self.selib.wait_until_element_is_visible(message_location,30)
             text_message =  self.selib.get_text(message_location)
             return     text_message
        
        def Get_email_verification_code_from_control (self):
             self.selib.wait_until_element_is_visible("//td/p[3]",30)
             email_code= self.selib.get_text("//td/p[3]")
             return     email_code
        


        def Do_Quick_quick_pick_for_all_boards(self):
                self.builtin.sleep(5)
                xpath="//fieldset[1]/*//button[contains(@data-pd,'quickpick')]"
                board_count=self.selib.get_webelements(xpath)
                board_count=len(board_count)
                for x in range(0,board_count):
                        xpath="//fieldset[1]/*//button[contains(@data-pd,'quickpickbutton_"+str(x)+"')]"
                        self.selib.click_element(xpath)
                        self.builtin.sleep(5)


        def Pick_DOB_from_calender(self,dob_loc,dob):
                self.selib.click_element(dob_loc)
                print(dob)
                dob_split = dob.split("/")
                Month = dob_split[0]
                print(Month)
                Date= dob_split[1]
                print(Date)
                birth_year= dob_split[2]
                print(birth_year)
                Current_years=self.selection_of_year(birth_year)
                for x in range(0,10):
                        
                        if (birth_year in Current_years):
                                self.selib.click_element("//div[@class='datepicker-years']/*//span[text()='"+birth_year+"']")
                                break
                        else:
                                self.selib.click_element("//div[@class='datepicker-years']/*//th[1]")
                                Current_years=self.selection_of_year(birth_year)               
                                

                self.selib.click_element("//div[@class='datepicker-months']/*//span[text()='"+Month+"']")
                self.builtin.sleep(2)
                self.selib.click_element("//div[@class='datepicker-days']/*//td[text()='"+Date+"']")

        def selection_of_year(self,birth_year):
                years_list=len(self.selib.get_webelements("//div[@class='datepicker-years']/*//span"))
                print("No of active years available in calender : ",years_list)
                Current_years=[]
                for x in range(0,int(years_list)):
                        x=x+1
                        year_available = self.selib.get_text("//div[@class='datepicker-years']/*//span["+str(x)+"]")
                        Current_years.append(year_available)
                print(Current_years)       
                return  Current_years

   
        
        def Do_quckpick_for_all_active_boards(self,no_boards):
                print("No of boards : ",no_boards)

        
        def convert_game_name_single_string(self,game_name):
                gname=game_name.lower()
                gname=gname.replace(" ","")
                print("Game Name is :",gname)
                if ("All or Nothing" in game_name):
                        gname='aon'
                
                if ("ikeno" in game_name or "iKeno" in game_name):
                        gname='keno'
                
                self.builtin.sleep(3)
                return gname


        # def import_data_from_xlsx(self,sheetname,xlsx_file):
        #         wb = load_workbook(filename=xlsx_file)

        #         # Select the active worksheet
        #         sheet = wb[sheetname]

        #         # Get dimensions of the sheet (rows and columns)
        #         max_row = sheet.max_row
        #         max_column = sheet.max_column

        #         # Accessing cell values
        #         print("Contents of the Excel file:")
        #         result_data =[]
        #         for row in range(1, max_row):
        #                 row_data = []
        #                 for column in range(1, max_column):
        #                         cell_value = sheet.cell(row=row+1, column=column+1).value
        #                         print(f"Row {row}, Column {column}: {cell_value}")
        #                         if isinstance(cell_value,float):
        #                                 data=int(cell_value)
        #                         else:
        #                                 data=str(cell_value)   
                        
        #                         print('data='+str(data))
        #                         data1=str(data)
        #                         row_data.append(data1)
        #                 result_data.append(row_data)
        #         # Accessing specific cells
        #         # For example, accessing cell A1
        #         cell_A1 = sheet['A1'].value
        #         print("\nValue in cell A1:", cell_A1)

        #         # Close the workbook
        #         wb.close()
        #         return  result_data
        
        def import_data_from_xlsx(self,sheetname,xlsx_file):

                # workbook object is created
                wb_obj = openpyxl.load_workbook(filename=xlsx_file)
                excel_data=[]
                sheet_obj = wb_obj[sheetname]
                max_col = sheet_obj.max_column
                max_rows = sheet_obj.max_row
                # Loop will print all columns name

                for j in range(2, max_rows + 1):
                        accessibility_row_data=[]
                        for i in range(1, max_col + 1):
                                cell_obj = sheet_obj.cell(row=j, column=i)
                                print(cell_obj.value)
                                accessibility_row_data.append(cell_obj.value)
                
                        excel_data.append(accessibility_row_data)

                wb_obj.close()
                return excel_data  
        				
        def export_data_from_xlsx(self,sheetname,xlsx_file,c1,c2,c3,c4,counter):
                print("vio",c1)
                print("Inapp",c2)
                print("Incomp",c3)
                print("Pass",c4)
                print("row value is",counter)

                wb_obj = openpyxl.load_workbook(filename=xlsx_file)

                sheet_obj = wb_obj[sheetname]
                max_col = 6
                max_rows = sheet_obj.max_row
                        # for counter in range(counter+2, max_rows + 1):
                        #         b1 = sheet_obj.cell(row = counter, column = max_col + 1) 
                        #         b2 = sheet_obj.cell(row = counter, column = max_col + 2) 
                        #         b3 = sheet_obj.cell(row = counter, column = max_col + 3) 
                        #         b4 = sheet_obj.cell(row = counter, column = max_col + 4) 
                        #         # writing values to cells 
                        #         b1.value = c1
                        #         b2.value = c2
                        #         b3.value = c3
                        #         b4.value = c4
                b1 = sheet_obj.cell(row = counter+2, column = max_col + 1) 
                b2 = sheet_obj.cell(row = counter+2, column = max_col + 2) 
                b3 = sheet_obj.cell(row = counter+2, column = max_col + 3) 
                b4 = sheet_obj.cell(row = counter+2, column = max_col + 4) 
                # writing values to cells 
                b1.value = c1
                b2.value = c2
                b3.value = c3
                b4.value = c4

                wb_obj.save(filename=xlsx_file)
        
        
        
                









        def readxl_new(self,sheetname,xlsx_file):
                newdir=os.curdir
                print(newdir)
                workbook = openpyxl.load_workbook(xlsx_file)
                sheet = workbook[sheetname]
                rowcount = sheet.max_row
                colcount = sheet.max_column
                print('rowcount'+str(rowcount))
                print('colcount'+str(colcount))
                
                result_data =[]
                
                for curr_row in range(1, rowcount, 1):
                
                        row_data = []
                for curr_col in range(1, colcount, 1):
                        # Read the data in the current cell
                        data = sheet.cell(curr_row+1, curr_col+1).value            
                        
                        if isinstance(data,float):
                                data=int(data)
                        else:
                                data=str(data)   
                        
                        print('data='+str(data))
                        data1=str(data)
                        row_data.append(data1)
                        
                result_data.append(row_data)
                
                return result_data  
        
        def Get_urls_from_excel(self,sheetname,xlsx_file):
                wb = load_workbook(filename=xlsx_file)

                # Select the active worksheet
                sheet = wb[sheetname]

                # Get dimensions of the sheet (rows and columns)
                max_row = sheet.max_row
                max_column = sheet.max_column

                # Accessing cell values
                print("Contents of the Excel file:")
                result_data =[]
                for row in range(1, max_row):
                        row_data = []
                        for column in range(1, max_column):
                                cell_value = sheet.cell(row=row+1, column=column+1).value
                                print(f"Row {row}, Column {column}: {cell_value}")
                                if isinstance(cell_value,float):
                                        data=int(cell_value)
                                else:
                                        data=str(cell_value)   
                        
                                print('data='+str(data))
                                data1=str(data)
                                row_data.append(data1)
                        result_data.append(row_data)
                # Accessing specific cells
                # For example, accessing cell A1
                cell_A1 = sheet['A1'].value
                print("\nValue in cell A1:", cell_A1)

                # Close the workbook
                wb.close()
                return  result_data
        def Create_Arrylist_with_output_values(self, page_url, page_title, Meta_description, Meta_keywords, H1_value, H2_value):
                print(page_url)
                result_data =[]
                for i in range(0, 6):
                        result_data.append(page_url)
                        result_data.append(page_title)
                        result_data.append(Meta_description)
                        result_data.append(Meta_keywords)
                        result_data.append(H1_value)
                        result_data.append(H2_value)

                global final_data
                final_data.append[result_data]
                return  final_data
        
        def fn(final_data=final_data):
                final_data.append(1)
                

        def write_test_results_in_excel(self,index_value,url='NA',page_title='NA',meta_description='NA',meta_keyword='NA',h1='NA',h2='NA'):
                print ("we are in writing methid")
                print (index_value)
                print (url)
                print (page_title)
                print (meta_description)
                print (meta_keyword)
                print (h1)
                print (h2)
                workbook = xlsxwriter.Workbook('SEO_Keyword.xlsx')
                data=[]
                data.append[index_value]
                data.append[url]
                data.append[page_title]
                data.append[meta_description]
                data.append[meta_keyword]
                data.append[h1]
                data.append[h2]
                print(data)
                # By default worksheet names in the spreadsheet will be 
                # Sheet1, Sheet2 etc., but we can also specify a name.
                worksheet = workbook.add_worksheet("My sheet")
                # Start from the first cell. Rows and
                # columns are zero indexed.
                row = 0
                col = index_value
                
                # Iterate over the data and write it out row by row.
                for index, url,page_title,meta_description,meta_keyword,h1,h2 in (data):
                        worksheet.write(row, col, index)
                        worksheet.write(row+1, col, url)
                        worksheet.write(row+2, col, page_title)
                        worksheet.write(row+3, col, meta_description)
                        worksheet.write(row+4, col, meta_keyword)
                        worksheet.write(row+5, col, h1)
                        worksheet.write(row+5, col, h2)
                        # col += 1
                
                workbook.close()




    
      





        
                  
